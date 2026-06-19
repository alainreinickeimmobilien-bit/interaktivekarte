#!/usr/bin/env python3
"""
Einmalige Migration: liest Grundstuecke_Uebersicht.xlsx und bereits_gefunden.json
ein und befüllt die neue SQLite-DB (properties + seen_urls).
Aufruf: python3 migrate_excel_to_db.py [pfad/zur/excel.xlsx]
"""
import json
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook

from app.db import SessionLocal, init_db
from app.models import Property, SeenUrl

GRUNDSTUECK_SYSTEM_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "grundstueck_system"
)
sys.path.insert(0, GRUNDSTUECK_SYSTEM_DIR)
from config import EXCEL_OUT, SEEN_FILE  # noqa: E402

RED_FILL_COLORS = {"FFCCCC", "00FFCCCC", "FFFFCCCC"}  # openpyxl liefert ARGB, unterschiedliche Alpha-Präfixe

COLS = [
    "datum", "plattform", "plz", "ort", "adresse",
    "groesse", "preis", "bplan", "anbieter", "link",
    "status", "propstack_id", "notizen", "haustyp",
]


def _cellval(row, idx):
    val = row[idx] if idx < len(row) else None
    return str(val).strip() if val not in (None, "") else ""


def migrate(excel_path: str):
    init_db()
    db = SessionLocal()

    wb = load_workbook(excel_path)
    ws = wb.active

    existing_links = {row.link for row in db.query(Property.link).all()}
    imported, skipped, expired = 0, 0, 0

    for row_idx in range(2, ws.max_row + 1):
        row = [ws.cell(row_idx, c).value for c in range(1, 15)]
        link = _cellval(row, 9)
        if not link or link in existing_links:
            skipped += 1
            continue

        fill = ws.cell(row_idx, 1).fill
        is_expired = bool(
            fill and fill.fgColor and fill.fgColor.rgb in RED_FILL_COLORS
        )
        if is_expired:
            expired += 1

        db.add(
            Property(
                datum=_cellval(row, 0),
                plattform=_cellval(row, 1),
                plz=_cellval(row, 2),
                ort=_cellval(row, 3),
                adresse=_cellval(row, 4),
                groesse=_cellval(row, 5),
                preis=_cellval(row, 6),
                bplan=_cellval(row, 7),
                anbieter=_cellval(row, 8),
                link=link,
                status=_cellval(row, 10) or "Neu",
                propstack_id=_cellval(row, 11),
                notizen=_cellval(row, 12),
                haustyp=_cellval(row, 13),
                is_expired=is_expired,
            )
        )
        existing_links.add(link)
        imported += 1

    db.commit()

    seen_count = 0
    if os.path.exists(SEEN_FILE):
        with open(SEEN_FILE, "r", encoding="utf-8") as f:
            seen_urls = set(json.load(f))
        existing_seen = {row.url for row in db.query(SeenUrl.url).all()}
        for url in seen_urls - existing_seen:
            db.add(SeenUrl(url=url))
            seen_count += 1
        db.commit()

    db.close()
    print(f"✅ {imported} Grundstücke importiert ({expired} davon als abgelaufen markiert)")
    print(f"   {skipped} übersprungen (bereits in DB)")
    print(f"   {seen_count} seen_urls aus {SEEN_FILE} übernommen")


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else EXCEL_OUT
    migrate(path)
