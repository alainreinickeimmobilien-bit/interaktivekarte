# ============================================================
# excel_handler.py – Excel-Übersicht lesen & schreiben
# ============================================================

import json
import os
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import EXCEL_OUT, SEEN_FILE


# Spaltenreihenfolge (muss zur Excel-Datei passen!)
COLUMNS = [
    "datum", "plattform", "plz", "ort", "adresse",
    "groesse", "preis", "bplan", "anbieter", "link",
    "status", "in_propstack", "notizen",
]

# Farben für Status
STATUS_FILLS = {
    "Neu":       PatternFill("solid", start_color="FFF2CC"),
    "Geprüft":   PatternFill("solid", start_color="D9EAD3"),
    "Import":    PatternFill("solid", start_color="B6D7A8"),
    "Abgelehnt": PatternFill("solid", start_color="F4CCCC"),
}
ALT_FILL   = PatternFill("solid", start_color="EBF5FB")
WHITE_FILL = PatternFill("solid", start_color="FFFFFF")

THIN   = Border(*[Side(style="thin", color="BFBFBF")] * 4)
FONT   = Font(name="Arial", size=9)


def _lade_gesehen() -> set:
    if os.path.exists(SEEN_FILE):
        with open(SEEN_FILE, "r", encoding="utf-8") as f:
            return set(json.load(f))
    return set()


def _speichere_gesehen(gesehen: set):
    with open(SEEN_FILE, "w", encoding="utf-8") as f:
        json.dump(list(gesehen), f, ensure_ascii=False)


def _wert(eintrag: dict, key: str) -> str:
    val = eintrag.get(key, "")
    return str(val).strip() if val else ""


def schreibe_in_excel(neue_eintraege: list) -> int:
    """
    Schreibt neue Grundstücke in die Excel-Übersicht.
    Skippt Duplikate (gleicher Link).
    Gibt Anzahl tatsächlich neu geschriebener Einträge zurück.
    """
    gesehen = _lade_gesehen()
    wb = load_workbook(EXCEL_OUT)
    ws = wb.active

    # Alle vorhandenen Links einlesen (Spalte J = Index 10)
    vorhandene_links = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[9]:  # Link-Spalte
            vorhandene_links.add(str(row[9]).strip())

    heute = date.today().strftime("%d.%m.%Y")
    neu_geschrieben = 0

    for eintrag in neue_eintraege:
        link = _wert(eintrag, "link")

        # Duplikatprüfung
        if link in vorhandene_links or link in gesehen:
            continue

        # Neue Zeile anhängen
        row_idx = ws.max_row + 1
        fill    = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL

        werte = [
            heute,
            _wert(eintrag, "plattform"),
            _wert(eintrag, "plz"),
            _wert(eintrag, "ort"),
            _wert(eintrag, "adresse"),
            _wert(eintrag, "groesse"),
            _wert(eintrag, "preis"),
            _wert(eintrag, "bplan"),
            _wert(eintrag, "anbieter"),
            link,
            "Neu",       # Status
            "nein",      # In Propstack
            _wert(eintrag, "notizen"),
        ]

        thin = Side(style="thin", color="BFBFBF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, wert in enumerate(werte, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=wert)
            cell.font      = Font(name="Arial", size=9)
            cell.fill      = fill
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 13))
            cell.border    = border

        vorhandene_links.add(link)
        gesehen.add(link)
        neu_geschrieben += 1

    wb.save(EXCEL_OUT)
    _speichere_gesehen(gesehen)
    return neu_geschrieben
