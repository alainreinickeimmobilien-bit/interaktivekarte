#!/usr/bin/env python3
"""
gs_update — Grundstücke Excel aktualisieren + abgelaufene Einträge rot markieren
Aufruf: python3 gs_update.py
"""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import requests
import openpyxl
from openpyxl.styles import PatternFill
from concurrent.futures import ThreadPoolExecutor
from config import EXCEL_OUT

HEADERS = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"}

RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

EXPIRED_KEYWORDS = [
    "nicht mehr verfügbar", "bereits verkauft", "angebot wurde beendet",
    "anzeige nicht mehr aktiv", "leider nicht mehr", "abgelaufen",
    "objekt nicht gefunden", "exposé nicht mehr verfügbar",
    "dieses inserat ist nicht mehr aktiv", "diese anzeige ist nicht mehr aktiv",
    "angebot nicht mehr vorhanden",
]


def check_link(entry):
    row_idx, link, ort, plz = entry
    try:
        r = requests.get(link, headers=HEADERS, timeout=10, allow_redirects=True)
        if r.status_code == 404:
            return (row_idx, link, ort, plz, 404, True)
        if r.status_code == 200:
            body = r.text.lower()
            if any(kw in body for kw in EXPIRED_KEYWORDS):
                return (row_idx, link, ort, plz, 200, True)
        return (row_idx, link, ort, plz, r.status_code, False)
    except Exception:
        return (row_idx, link, ort, plz, "ERR", False)


def markiere_abgelaufen():
    wb = openpyxl.load_workbook(EXCEL_OUT)
    ws = wb.active

    to_check = []
    for i in range(2, ws.max_row + 1):
        link = ws.cell(i, 10).value
        if link:
            to_check.append((i, link, ws.cell(i, 4).value, ws.cell(i, 3).value))

    print(f"🔗 Prüfe {len(to_check)} Links auf Gültigkeit ...")
    with ThreadPoolExecutor(max_workers=15) as exe:
        results = list(exe.map(check_link, to_check))

    expired_rows = [r[0] for r in results if r[5]]

    if not expired_rows:
        print("✅ Keine abgelaufenen Einträge gefunden.")
        return 0

    for row_idx in expired_rows:
        ort = ws.cell(row_idx, 4).value
        plz = ws.cell(row_idx, 3).value
        print(f"  🔴  Markiere abgelaufen: {ort} ({plz}) — Zeile {row_idx}")
        for col in range(1, ws.max_column + 1):
            ws.cell(row_idx, col).fill = RED_FILL

    wb.save(EXCEL_OUT)
    print(f"\n✅ {len(expired_rows)} abgelaufene Einträge rot markiert.")
    return len(expired_rows)


if __name__ == "__main__":
    # 1. Neue Grundstücke scrapen (main.py-Logik)
    print("=" * 60)
    print("  SCHRITT 1: Neue Grundstücke suchen")
    print("=" * 60)
    try:
        import main as m
        m.main()
    except Exception as e:
        print(f"⚠️  Scraper-Fehler: {e}")

    # 2. Abgelaufene Einträge prüfen und rot markieren
    print("\n" + "=" * 60)
    print("  SCHRITT 2: Abgelaufene Einträge markieren")
    print("=" * 60)
    markiere_abgelaufen()
