#!/usr/bin/env python3
# ============================================================
# preisliste_extraktion.py – massa haus Preisbücher auslesen
#
# Liest aus allen 7 Preisbuch-PDFs je Haustyp:
#   1. Ausbauhauspreis (Standardgrundriss)
#   2a. Kompaktanlage Luft-Wasser-WP, Preise inkl. Standard-Bodenplatte,
#       Heizung inkl. Montage  (ComfortStyle / LifeStyle / Trend / TwinStyle / UrbanStyle)
#       → bei n.m. Fallback auf Zeile darunter: Innenaufstellung (FamilyStyle)
#   2b. Fast-Fertig-Option  (UnitStyle, direkt auf der Detailseite)
#
# Ausgabe: SQLite-Datenbank + Excel
#   - Preisliste_massa_haus.db   (Datenbank, immer aktuell)
#   - Preisliste_massa_haus.xlsx (Excel zum Anschauen)
#
# Aufruf: python3 preisliste_extraktion.py
# Bei Preisänderung: Ordnerpfad anpassen + erneut ausführen → DB aktualisiert
# ============================================================

import os
import re
import sqlite3
from datetime import datetime
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Pfade ────────────────────────────────────────────────────
PDF_ORDNER = "/Users/sophiakroehner/Desktop/Massa/2_ZUM KALKULIEREN/11.2025"
EXCEL_OUT  = "/Users/sophiakroehner/Desktop/Massa/2_ZUM KALKULIEREN/Preisliste_massa_haus.xlsx"
DB_OUT     = "/Users/sophiakroehner/Desktop/Massa/2_ZUM KALKULIEREN/Preisliste_massa_haus.db"

PREISBUECHER = [
    "2025_11 Preisbuch ComfortStyle.pdf",
    "2025_11 Preisbuch FamilyStyle.pdf",
    "2025_11 Preisbuch LifeStyle.pdf",
    "2025_11 Preisbuch Trend.pdf",
    "2025_11 Preisbuch TwinStyle.pdf",
    "2025_11 Preisbuch UnitStyle.pdf",
    "2025_11 Preisbuch UrbanStyle.pdf",
]

# Hausreihen, die Fast-Fertig-Option statt WP verwenden
FAST_FERTIG_REIHEN = {"UnitStyle"}


# ── Hilfsfunktionen ──────────────────────────────────────────

def _preis_bereinigen(text: str) -> str:
    """
    Bereinigt PDF-Extraktionsfehler in Preisen.
    '1 30.610,- €' → '130.610 €'
    '150.999,- €'  → '150.999 €'
    'n.m.'         → 'n.m.'
    """
    if not text:
        return ""
    t = str(text).strip()
    if t.lower() in ("n.m.", "nm", ""):
        return "n.m."
    # Leerzeichen zwischen Ziffern entfernen (PDF-Artefakt: "1 30.610" → "130.610")
    t = re.sub(r"(\d)\s+(\d)", r"\1\2", t)
    # Komma+Bindestrich am Ende entfernen
    t = re.sub(r",\s*-", "", t)
    # Zahl extrahieren
    m = re.search(r"[\d\.]+", t)
    if m:
        zahl = m.group().replace(".", "")
        try:
            return f"{int(zahl):,} €".replace(",", ".")
        except ValueError:
            pass
    return t


def _preis_zu_zahl(preis_str: str):
    """'150.999 €' → 150999.0  |  'n.m.' oder '' → None"""
    if not preis_str or preis_str.strip().lower() in ("n.m.", ""):
        return None
    clean = preis_str.replace(".", "").replace(",", ".").replace("€", "").strip()
    try:
        return float(clean)
    except ValueError:
        return None


def _preis_aus_zeile(row: list, col_idx) -> str:
    """
    Extrahiert den Preis aus einer Tabellenzeile.
    Strategie A: direkter Spaltenindex (aus Kopfzeile ermittelt).
    Strategie B: Preiswerte aus der Zeile sammeln, Größenfelder ausschließen,
                 dann Index 1 nehmen (Material=0, Heizung inkl. Montage=1).
    """
    # Strategie A
    if col_idx is not None and col_idx < len(row):
        val = row[col_idx]
        if val and str(val).strip():
            return _preis_bereinigen(str(val))

    # Strategie B
    preis_werte = []
    for c in row:
        if not c:
            continue
        w = str(c).strip()
        if not w:
            continue
        if any(x in w for x in ("m²", "m2", "Nutzfläche", "max.", "beheizt")):
            continue
        if re.search(r"\d{3,}", w) and re.search(r"[€,\.]", w):
            preis_werte.append(w)
        elif w.lower() in ("n.m.", "nm"):
            preis_werte.append("n.m.")

    if len(preis_werte) >= 2:
        return _preis_bereinigen(preis_werte[1])
    elif preis_werte:
        return _preis_bereinigen(preis_werte[0])
    return ""


# ── Extraktionsfunktionen ────────────────────────────────────

def _extrahiere_ausbauhauspreis(page) -> str:
    """Zeile 'Ausbauhaus (Standardgrundriss)' → letzte nicht-leere Zelle."""
    for table in page.extract_tables():
        for row in table:
            zellen = [c for c in row if c]
            if not zellen:
                continue
            if "Ausbauhaus" in zellen[0] and "Standardgrundriss" in zellen[0]:
                return _preis_bereinigen(zellen[-1])
    return ""


def _extrahiere_fast_fertig_preis(page) -> str:
    """
    UnitStyle-Detailseite: Zeile 'Fast-Fertig Option' → letzte nicht-leere Zelle.
    Sitzt auf der GLEICHEN Seite wie der Ausbauhauspreis.
    """
    for table in page.extract_tables():
        for row in table:
            zellen = [c for c in row if c]
            if not zellen:
                continue
            erste = str(zellen[0]).strip()
            if "Fast-Fertig" in erste and "Option" in erste:
                return _preis_bereinigen(zellen[-1])
    return ""


def _extrahiere_fast_fertig_ausbaupaket(page) -> str:
    """
    Ausbaupakete-Seite (Register 3, Seite i+2 nach der Detailseite):
    Findet die Spalte 'Fast-Fertig Option' (ganz rechts) und gibt
    den ersten Preis in dieser Spalte zurück (erste Datenzeile = Dämmpaket).

    Gilt für: ComfortStyle, FamilyStyle, LifeStyle, Trend, TwinStyle, UrbanStyle.
    """
    for table in page.extract_tables():
        # Schritt 1: Spaltenindex UND Zeilenindex der Fast-Fertig-Option-Kopfzeile
        ff_col       = None
        ff_header_row = None
        for row_idx, row in enumerate(table):
            for col_idx, cell in enumerate(row):
                if cell and "Fast-Fertig" in str(cell) and "Option" in str(cell):
                    ff_col        = col_idx
                    ff_header_row = row_idx
                    break
            if ff_col is not None:
                break

        if ff_col is None:
            continue

        # Schritt 2: Ersten echten Preiswert NUR in Zeilen NACH der Kopfzeile
        for row in table[ff_header_row + 1:]:
            if ff_col >= len(row):
                continue
            val = row[ff_col]
            if not val:
                continue
            val_str   = str(val).strip()
            bereinigt = _preis_bereinigen(val_str)
            # Nur echte Preise zurückgeben (enden mit '€')
            if bereinigt and bereinigt.endswith("€"):
                return bereinigt

    return ""


def _extrahiere_kompaktanlage_preis(page) -> str:
    """
    Technikpaket-Seite: Zeile 'Kompaktanlage Luft-Wasser-Wärmepumpe'
    → Spalte 'Preise inkl. Standard-Bodenplatte, Heizung inkl. Montage'.

    Tabellenstruktur (0-basiert):
      0  = Heizungspaket (rotierter Text)
      1  = Bezeichnung
      2  = Größenbereich (kein Preis!)
      3  = Standard-Bodenplatte, Material
      4  = Standard-Bodenplatte, Heizung inkl. Montage  ← GESUCHTER PREIS
      5  = Standard-Keller, Material
      6  = Standard-Keller, Heizung inkl. Montage

    FamilyStyle-Fallback:
      Wenn die Kompaktanlage-Zeile 'n.m.' liefert, wird die nächste Zeile
      (Innenaufstellung Heizungsanlage im EG oder KG) verwendet.
    """
    for table in page.extract_tables():
        # Schritt 1: Spaltenindex aus Kopfzeile ermitteln
        heizung_col = None
        for row in table:
            zellen = [str(c) if c else "" for c in row]
            roh = " ".join(zellen)
            if "Heizung" in roh and "inkl. Montage" in roh and "Material" in roh:
                montage_idx = [i for i, c in enumerate(zellen) if "Heizung" in c and "inkl" in c]
                if montage_idx:
                    heizung_col = montage_idx[0]
                break

        # Schritt 2: Kompaktanlage-Zeile finden
        for row_idx, row in enumerate(table):
            roh = " ".join(str(c) for c in row if c)
            if "Kompaktanlage" not in roh or "Wärmepumpe" not in roh:
                continue

            preis = _preis_aus_zeile(row, heizung_col)

            # FamilyStyle-Fallback: bei n.m. nächste Zeile probieren
            # (= Innenaufstellung Heizungsanlage im EG oder KG)
            if preis == "n.m." and row_idx + 1 < len(table):
                naechste = table[row_idx + 1]
                preis_naechste = _preis_aus_zeile(naechste, heizung_col)
                if preis_naechste and preis_naechste != "n.m.":
                    return preis_naechste

            if preis:
                return preis

    return ""


def _ist_haustyp_detailseite(page_text: str, hausreihe: str) -> str:
    """
    Prüft ob die Seite eine Detailseite für einen Haustyp ist.
    Gibt den vollständigen Haustyp-Namen zurück (inkl. .EH/.MH für UrbanStyle).
    """
    m = re.search(
        rf"Preisliste Haustyp\s+({re.escape(hausreihe)}\s+[\d\.]+\s*[A-Z\.]+)",
        page_text
    )
    if m:
        return m.group(1).strip()
    return ""


# ── PDF-Extraktion ───────────────────────────────────────────

def extrahiere_pdf(pdf_pfad: str, hausreihe: str) -> list:
    """
    Liest ein Preisbuch-PDF und gibt eine Liste mit Dicts zurück.
    """
    ergebnisse = []
    ist_fast_fertig = hausreihe in FAST_FERTIG_REIHEN
    print(f"\n  📖 Lese {os.path.basename(pdf_pfad)}...")

    with pdfplumber.open(pdf_pfad) as pdf:
        n = len(pdf.pages)
        i = 0
        while i < n:
            page     = pdf.pages[i]
            text     = page.extract_text() or ""
            haustyp  = _ist_haustyp_detailseite(text, hausreihe)

            if not haustyp:
                i += 1
                continue

            ausbauhaus  = _extrahiere_ausbauhauspreis(page)

            if not ausbauhaus:
                # Übersichtsseite (mehrere Häuser), kein eigener Eintrag
                i += 1
                continue

            kompaktanlage = ""
            fast_fertig   = ""

            if ist_fast_fertig:
                # UnitStyle: Fast-Fertig auf der GLEICHEN Seite
                fast_fertig = _extrahiere_fast_fertig_preis(page)
            else:
                # Alle anderen: Technikpaket auf Seite i+1
                if i + 1 < n:
                    naechste      = pdf.pages[i + 1]
                    naechste_text = naechste.extract_text() or ""
                    if "Kompaktanlage" in naechste_text or "Wärmepumpe" in naechste_text:
                        kompaktanlage = _extrahiere_kompaktanlage_preis(naechste)
                # Fast-Fertig-Option auf Ausbaupakete-Seite (i+2)
                if i + 2 < n:
                    ausbaupakete      = pdf.pages[i + 2]
                    ausbaupakete_text = ausbaupakete.extract_text() or ""
                    if "Fast-Fertig" in ausbaupakete_text:
                        fast_fertig = _extrahiere_fast_fertig_ausbaupaket(ausbaupakete)

            ergebnisse.append({
                "hausreihe":    hausreihe,
                "haustyp":      haustyp,
                "ausbauhaus":   ausbauhaus,
                "kompaktanlage": kompaktanlage,
                "fast_fertig":  fast_fertig,
            })

            label2 = f"Fast-Fertig={fast_fertig}" if ist_fast_fertig else f"WP={kompaktanlage}"
            print(f"    ✓ {haustyp}: Ausbauhaus={ausbauhaus} | {label2}")
            i += 2  # Detailseite überspringen (+ ggf. Technikpaket)

    return ergebnisse


# ── Datenbankausgabe ─────────────────────────────────────────

def schreibe_datenbank(alle_daten: list, stand: str):
    """
    Schreibt/aktualisiert die Preisdaten in der SQLite-Datenbank.
    Alte Stände bleiben erhalten → Preisentwicklung nachvollziehbar.
    """
    con = sqlite3.connect(DB_OUT)
    cur = con.cursor()

    # Tabelle anlegen
    cur.executescript("""
        CREATE TABLE IF NOT EXISTS preise (
            id                    INTEGER PRIMARY KEY AUTOINCREMENT,
            hausreihe             TEXT    NOT NULL,
            haustyp               TEXT    NOT NULL,
            stand                 TEXT    NOT NULL,
            ausbauhaus_text       TEXT,
            ausbauhaus_euro       REAL,
            kompaktanlage_text    TEXT,
            kompaktanlage_euro    REAL,
            fast_fertig_text      TEXT,
            fast_fertig_euro      REAL,
            aktualisiert          TEXT    NOT NULL,
            UNIQUE(haustyp, stand)
        );
        CREATE INDEX IF NOT EXISTS idx_haustyp ON preise(haustyp);
        CREATE INDEX IF NOT EXISTS idx_stand   ON preise(stand);
    """)

    # Spalten nachrüsten falls DB aus alter Version stammt
    vorhandene = {r[1] for r in cur.execute("PRAGMA table_info(preise)")}
    for spalte, typ in [("fast_fertig_text", "TEXT"), ("fast_fertig_euro", "REAL")]:
        if spalte not in vorhandene:
            cur.execute(f"ALTER TABLE preise ADD COLUMN {spalte} {typ}")

    con.commit()

    jetzt      = datetime.now().strftime("%Y-%m-%d %H:%M")
    eingefuegt = 0
    aktualisiert_n = 0

    for e in alle_daten:
        existing = cur.execute(
            "SELECT id FROM preise WHERE haustyp = ? AND stand = ?",
            (e["haustyp"], stand)
        ).fetchone()

        felder = (
            e["ausbauhaus"],          _preis_zu_zahl(e["ausbauhaus"]),
            e["kompaktanlage"],       _preis_zu_zahl(e["kompaktanlage"]),
            e["fast_fertig"],         _preis_zu_zahl(e["fast_fertig"]),
            jetzt,
        )

        if existing:
            cur.execute("""
                UPDATE preise SET
                    hausreihe=?,
                    ausbauhaus_text=?,   ausbauhaus_euro=?,
                    kompaktanlage_text=?, kompaktanlage_euro=?,
                    fast_fertig_text=?,  fast_fertig_euro=?,
                    aktualisiert=?
                WHERE haustyp=? AND stand=?
            """, (e["hausreihe"], *felder, e["haustyp"], stand))
            aktualisiert_n += 1
        else:
            cur.execute("""
                INSERT INTO preise
                    (hausreihe, haustyp, stand,
                     ausbauhaus_text, ausbauhaus_euro,
                     kompaktanlage_text, kompaktanlage_euro,
                     fast_fertig_text, fast_fertig_euro,
                     aktualisiert)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (e["hausreihe"], e["haustyp"], stand, *felder))
            eingefuegt += 1

    con.commit()
    con.close()
    print(f"  💾 DB: {eingefuegt} neu, {aktualisiert_n} aktualisiert → {DB_OUT}")


# ── Excel-Ausgabe ────────────────────────────────────────────

def schreibe_excel(alle_daten: list, stand: str):
    """Schreibt die extrahierten Preise in eine formatierte Excel-Datei."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Preisliste massa haus"

    DUNKELROT  = "8B0000"
    HELLGRAU   = "F2F2F2"
    WEISS      = "FFFFFF"
    DUNKELGRAU = "404040"

    header_font = Font(name="Calibri", bold=True, color=WEISS, size=11)
    header_fill = PatternFill("solid", fgColor=DUNKELROT)
    gruppe_font = Font(name="Calibri", bold=True, color=WEISS, size=10)
    gruppe_fill = PatternFill("solid", fgColor=DUNKELGRAU)
    normal_font = Font(name="Calibri", size=10)
    grau_fill   = PatternFill("solid", fgColor=HELLGRAU)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    center = Alignment(horizontal="center", vertical="center")
    links  = Alignment(horizontal="left",   vertical="center")

    # Kopfzeile
    headers = [
        "Hausreihe",
        "Haustyp",
        "Ausbauhaus\n(Standardgrundriss)",
        "LW-Wärmepumpe Kompaktanlage\nPreise inkl. Standard-Bodenplatte\n+ Heizung inkl. Montage\n(bei FamilyStyle: Innenaufstellung)",
        "Fast-Fertig-Option\n(alle Haustypen)",
    ]
    ws.row_dimensions[1].height = 55
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.border    = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Daten
    aktuelle_reihe = None
    zeile = 2
    alle_daten.sort(key=lambda x: (x["hausreihe"], x["haustyp"]))

    for e in alle_daten:
        if e["hausreihe"] != aktuelle_reihe:
            aktuelle_reihe = e["hausreihe"]
            ws.merge_cells(f"A{zeile}:E{zeile}")
            c = ws.cell(row=zeile, column=1, value=f"  {aktuelle_reihe}")
            c.font      = gruppe_font
            c.fill      = gruppe_fill
            c.border    = thin_border
            c.alignment = links
            ws.row_dimensions[zeile].height = 20
            zeile += 1

        fill  = grau_fill if zeile % 2 == 0 else None
        daten = [
            e["hausreihe"],
            e["haustyp"],
            e["ausbauhaus"],
            e["kompaktanlage"],
            e["fast_fertig"],
        ]
        for col, wert in enumerate(daten, 1):
            c = ws.cell(row=zeile, column=col, value=wert)
            c.font      = normal_font
            c.border    = thin_border
            c.alignment = center if col >= 3 else links
            if fill:
                c.fill = fill
        ws.row_dimensions[zeile].height = 18
        zeile += 1

    # Spaltenbreiten
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 32
    ws.column_dimensions["E"].width = 22

    ws.auto_filter.ref = f"A1:E{zeile-1}"
    ws.freeze_panes    = "A2"

    ws.cell(row=zeile + 1, column=1,
            value=f"Stand: {stand} | Alle Preise inkl. 19% MwSt. | Quelle: massa haus Preisbücher"
    ).font = Font(name="Calibri", italic=True, color="808080", size=9)

    wb.save(EXCEL_OUT)
    print(f"  💾 Excel gespeichert: {EXCEL_OUT}")


# ── Hauptprogramm ────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  MASSA HAUS PREISLISTEN-EXTRAKTION")
    print("=" * 60)

    alle_daten = []

    for dateiname in PREISBUECHER:
        pfad = os.path.join(PDF_ORDNER, dateiname)
        if not os.path.exists(pfad):
            print(f"  ⚠️  Nicht gefunden: {dateiname}")
            continue
        m = re.search(r"Preisbuch\s+(.+?)\.pdf", dateiname, re.IGNORECASE)
        hausreihe = m.group(1) if m else dateiname
        alle_daten.extend(extrahiere_pdf(pfad, hausreihe))

    stand = os.path.basename(PDF_ORDNER)  # z.B. "11.2025"

    print(f"\n{'=' * 60}")
    print(f"  {len(alle_daten)} Haustypen extrahiert  |  Stand: {stand}")

    if alle_daten:
        schreibe_datenbank(alle_daten, stand)
        schreibe_excel(alle_daten, stand)
        print(f"  ✅ FERTIG")
    else:
        print(f"  ❌ Keine Daten gefunden")

    print("=" * 60)


if __name__ == "__main__":
    main()
