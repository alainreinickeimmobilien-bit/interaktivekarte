#!/usr/bin/env python3
# ============================================================
# propstack_import.py – Markierte Grundstücke → Propstack CRM
#
# Workflow:
#   1. Excel öffnen (Grundstuecke_Uebersicht.xlsx)
#   2. Alle Zeilen mit Status = "Import" einlesen
#   3. Je Grundstück EINE Anzeige anlegen:
#      Gesamtpreis = Grundstück + Ausbauhaus + Technikpaket (LW-WP)
#   4. Status in Excel auf "In Propstack" setzen
#
# Aufruf: python3 propstack_import.py
# ============================================================

import os
import re
import sqlite3
import requests
from datetime import date
from openpyxl import load_workbook

import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import PROPSTACK_API_KEY, PROPSTACK_BASE, EXCEL_OUT

DB_PREISE = os.environ.get(
    "DB_PREISE", "/Users/sophiakroehner/Desktop/Massa/2_ZUM KALKULIEREN/Preisliste_massa_haus.db"
)

HEADERS = {
    "X-API-KEY": PROPSTACK_API_KEY,
    "Content-Type": "application/json",
    "Accept": "application/json",
}
HEADERS_UPLOAD = {"X-API-KEY": PROPSTACK_API_KEY}

BROKER_ID = 353664  # Sophia Kröhner

IMG_BASISORDNER = os.environ.get(
    "IMG_BASISORDNER", os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
)

# Spalten-Mapping (1-basiert)
COL = {
    "datum":        1,
    "plattform":    2,
    "plz":          3,
    "ort":          4,
    "adresse":      5,
    "groesse":      6,
    "preis":        7,
    "bplan":        8,
    "anbieter":     9,
    "link":         10,
    "status":       11,
    "in_propstack": 12,
    "notizen":      13,
    "haustyp":      14,
}

# Status-Werte die als "zu importieren" gelten
IMPORT_STATI = {"import", "import fehlt haustyp"}


# ── Preise aus DB ─────────────────────────────────────────────

def _lade_hauspreise(haustyp: str) -> dict:
    """Lädt Ausbauhaus- und Technikpaket-Preis aus der Preisliste-DB."""
    if not os.path.exists(DB_PREISE):
        print(f"  ⚠️  Preisliste-DB nicht gefunden: {DB_PREISE}")
        return {}
    con = sqlite3.connect(DB_PREISE)
    cur = con.cursor()
    row = cur.execute("""
        SELECT ausbauhaus_euro, kompaktanlage_euro
        FROM preise
        WHERE haustyp = ?
        ORDER BY stand DESC LIMIT 1
    """, (haustyp,)).fetchone()
    con.close()
    if not row:
        print(f"  ⚠️  Haustyp '{haustyp}' nicht in Preisliste gefunden.")
        return {}
    return {
        "ausbauhaus":    row[0] or 0.0,
        "technikpaket":  row[1] or 0.0,
    }


# ── Preisformatierung ─────────────────────────────────────────

def _fmt(euro: float) -> str:
    """123456.0 → '123.456 €'"""
    return f"{int(euro):,} €".replace(",", ".")


def _parse_grundstueck_preis(preis_str) -> float:
    """'38.000 €' oder 38000 → 38000.0"""
    if not preis_str:
        return 0.0
    clean = str(preis_str).replace(".", "").replace(",", ".").replace("€", "").strip()
    try:
        return float(clean)
    except ValueError:
        return 0.0


# ── Lage-Beschreibung ─────────────────────────────────────────

LAGEBESCHREIBUNGEN = {
    # Mittelsachsen
    "09577": "Niederwiesa liegt im Landkreis Mittelsachsen, eingebettet in die grüne Hügellandschaft zwischen Chemnitz und Frankenberg. Die Gemeinde überzeugt mit ruhiger Wohnlage und naturnahem Umfeld – ideal für Familien, die Erholung suchen, ohne auf gute Anbindung zu verzichten. Chemnitz ist in etwa 20 Minuten erreichbar.",
    "09661": "Hainichen liegt im Herzen des Landkreises Mittelsachsen. Die Kleinstadt bietet eine gewachsene Infrastruktur mit Schulen, Einkaufsmöglichkeiten und guter Verkehrsanbindung an Chemnitz und Freiberg. Die umliegende Natur lädt zum Erholen ein.",
    "04720": "Döbeln und Umgebung liegen im Landkreis Mittelsachsen an der Freiberger Mulde. Die Region bietet ein breites Infrastrukturangebot und ist gut an die Autobahn sowie die Bahnstrecke Leipzig–Dresden angebunden.",
    "04736": "Waldheim liegt im Landkreis Mittelsachsen an der Zschopau – eine ruhige Kleinstadt mit solider Infrastruktur und guter Anbindung an Döbeln und Chemnitz.",
    "04741": "Roßwein liegt im Landkreis Mittelsachsen an der Freiberger Mulde mit guter Anbindung nach Leipzig und Chemnitz.",
    "09131": "Chemnitz ist die drittgrößte Stadt Sachsens und bietet alle Vorteile einer modernen Großstadt: vielfältige Einkaufsmöglichkeiten, Schulen aller Art, Kulturangebote und eine ausgezeichnete Verkehrsanbindung.",
    "09114": "Chemnitz ist die drittgrößte Stadt Sachsens und bietet alle Vorteile einer modernen Großstadt: vielfältige Einkaufsmöglichkeiten, Schulen aller Art, Kulturangebote und eine ausgezeichnete Verkehrsanbindung.",
    "09123": "Chemnitz-Röhrsdorf liegt im Südwesten von Chemnitz und verbindet ruhige Wohnlage mit guter Stadtanbindung. Die naturnahe Umgebung und die Nähe zur Autobahn machen den Standort besonders attraktiv für Familien.",
    "09228": "Chemnitz-Klaffenbach liegt im südlichen Stadtgebiet von Chemnitz in einer grünen, ruhigen Lage. Die Region bietet Natur pur bei gleichzeitiger Nähe zur Großstadt.",
    "09244": "Lichtenau liegt im Landkreis Mittelsachsen südwestlich von Chemnitz. Die Gemeinde punktet mit ruhiger Lage, Naturnähe und guter Erreichbarkeit der Stadt.",
    "09306": "Rochlitz liegt im Landkreis Mittelsachsen an der Zwickauer Mulde. Die historische Kleinstadt bietet solide Infrastruktur und eine malerische Umgebung.",
    "09326": "Geringswalde liegt im Landkreis Mittelsachsen und bietet eine ruhige, ländliche Wohnatmosphäre mit Anschluss an das überregionale Straßennetz.",
    # Erzgebirge
    "09405": "Zschopau liegt im Erzgebirgskreis im Zschopautal. Die Stadt bietet gute Infrastruktur, Schulen und ist gut an Chemnitz angebunden.",
    "09430": "Annaberg-Buchholz ist die Kreisstadt des Erzgebirgskreises und bietet als regionales Zentrum eine vollständige Infrastruktur mit Schulen, Einkauf und Gesundheitsversorgung.",
    "09432": "Großolbersdorf liegt im Erzgebirgskreis in einer naturnahen, ruhigen Höhenlage mit Anschluss an Zschopau und Chemnitz.",
    "09434": "Zschopau / Krumhermersdorf liegt im Erzgebirgskreis in einer ruhigen, naturnahen Lage im Zschopautal.",
    "09437": "Hopfgarten liegt im Erzgebirgskreis, eingebettet in die typische Erzgebirgslandschaft mit ruhiger Wohnlage und guter Anbindung.",
    "09439": "Amtsberg liegt im Erzgebirgskreis in einer ruhigen Erzgebirgslage mit guten Verbindungen nach Chemnitz und Annaberg-Buchholz.",
    "09456": "Annaberg-Buchholz ist die Kreisstadt des Erzgebirgskreises mit vollständiger Infrastruktur und guter Anbindung.",
    "09496": "Marienberg liegt im Erzgebirgskreis und bietet als Kleinstadt ein gutes Infrastrukturangebot inmitten der Erzgebirgslandschaft.",
    "09509": "Großrückerswalde liegt im Erzgebirgskreis in einer ruhigen Lage mit Anschluss an Marienberg und Chemnitz.",
    "09518": "Großrückerswalde / Rübenau liegt im Erzgebirgskreis, umgeben von Natur und ruhiger Wohnlage.",
    "09526": "Olbernhau liegt im Erzgebirgskreis im Flöhatal. Die Kleinstadt bietet Infrastruktur und eine natürliche, waldreiche Umgebung.",
    "09548": "Deutschneudorf liegt im Erzgebirgskreis in idyllischer Höhenlage nahe dem Erzgebirgskamm – ruhige Natur pur, ideal für Familien.",
    "09557": "Flöha liegt im Landkreis Mittelsachsen an der Flöha und bietet gute Anbindung an Chemnitz und Marienberg.",
    "09569": "Oederan liegt im Landkreis Mittelsachsen westlich von Freiberg. Die Kleinstadt bietet solide Infrastruktur und ruhige Wohnlagen.",
    "09573": "Augustusburg liegt im Erzgebirgskreis hoch über dem Zschopautal und ist bekannt für sein imposantes Schloss. Die Lage bietet Panoramablicke und ruhige Wohnatmosphäre.",
    "09579": "Hilbersdorf / Hetzdorf liegt im Landkreis Mittelsachsen in ruhiger, ländlicher Umgebung mit guter Anbindung an Chemnitz.",
    "09599": "Freiberg ist die Kreisstadt des Landkreises Mittelsachsen, Universitätsstadt und bietet eine vollständige Infrastruktur sowie hervorragende Anbindung.",
    "09600": "Weißenborn liegt im Landkreis Mittelsachsen nahe Freiberg in ruhiger, ländlicher Lage.",
    "09603": "Großschirma liegt im Landkreis Mittelsachsen zwischen Freiberg und Großenhain mit guter Verkehrsanbindung.",
    "09618": "Brand-Erbisdorf liegt im Landkreis Mittelsachsen südlich von Freiberg, eingebettet in die Erzgebirgslandschaft.",
    "09619": "Mulda liegt im Landkreis Mittelsachsen in ruhiger Lage mit Anschluss an Freiberg und Chemnitz.",
    "09623": "Frauenstein liegt im Erzgebirgskreis in reizvoller Höhenlage mit natürlichem Umfeld.",
    "09627": "Bobritzsch-Hilbersdorf liegt im Landkreis Mittelsachsen in ruhiger, ländlicher Lage nahe Freiberg.",
    "09629": "Bieberstein / Reinsberg liegt im Landkreis Mittelsachsen zwischen Freiberg und Döbeln in einer ruhigen, naturnahen Wohnlage mit Blick ins Grüne.",
    "09633": "Halsbrücke liegt im Landkreis Mittelsachsen nahe Freiberg in ruhiger Lage.",
    "09638": "Lichtenberg/Erzgeb. liegt im Erzgebirgskreis in einer ruhigen Erzgebirgslage.",
    "09648": "Mittweida liegt im Landkreis Mittelsachsen an der Zschopau. Die Hochschulstadt bietet lebendige Infrastruktur und gute Verkehrsanbindung.",
    "09669": "Striegistal liegt im Landkreis Mittelsachsen in einer ruhigen, ländlichen Lage zwischen Hainichen und Döbeln.",
    # Leipzig-Umland
    "04749": "Ostrau liegt im Landkreis Mittelsachsen in ruhiger ländlicher Lage mit Anschluss an Döbeln und Leipzig.",
}

# Ortsbasierte Überschreibungen (wenn Ort nicht zum PLZ passt)
LAGE_ORT_OVERRIDE = {
    "Waldheim": "Waldheim liegt im Landkreis Mittelsachsen an der Zschopau – eine ruhige Kleinstadt mit solider Infrastruktur und guter Anbindung an Döbeln und Chemnitz.",
    "Großweitzschen": "Großweitzschen liegt im Landkreis Mittelsachsen in ruhiger, ländlicher Lage nahe Döbeln – ideal für Familien, die Natur und gute Erreichbarkeit verbinden möchten.",
}

def _lage_beschreibung(plz: str, ort: str) -> str:
    """Gibt eine Lagebeschreibung für PLZ/Ort zurück (kein Straßenname!)."""
    if ort and ort in LAGE_ORT_OVERRIDE:
        return LAGE_ORT_OVERRIDE[ort]
    if plz and plz in LAGEBESCHREIBUNGEN:
        return LAGEBESCHREIBUNGEN[plz]
    if ort:
        return (
            f"{ort} bietet eine ruhige, naturnahe Wohnlage in Sachsen mit guter "
            f"Anbindung an die umliegenden Städte. Die Region eignet sich besonders "
            f"für Familien, die Erholung und solide Infrastruktur verbinden möchten."
        )
    return ""


# ── Anzeigentexte ─────────────────────────────────────────────

def _titel(ort: str, haustyp: str) -> str:
    return f"Grundstück mit massa haus {haustyp} – technisch fertig in {ort}"


def _beschreibung(grundstueck: float, hauspreise: dict,
                  groesse: str, plz: str, ort: str, haustyp: str,
                  anbieter: str) -> tuple:
    """
    Erstellt (description_note, location_note).
    Gesamtpaket: Grundstück + Ausbauhaus + Technikpaket
    """
    ab = hauspreise.get("ausbauhaus", 0.0)
    tp = hauspreise.get("technikpaket", 0.0)
    gs = grundstueck
    gesamt = gs + ab + tp

    kerntext = (
        "Grundstück mit technisch-fertigem massa haus\n\n"
        "Dieses Angebot kombiniert ein Baugrundstück mit dem massa haus als Ausbauhaus "
        "inkl. vollständigem Technikpaket – alles aus einer Hand.\n\n"
        "Im Technikpaket enthalten:\n"
        "✓ Standard-Bodenplatte\n"
        "✓ Luft-Wasser-Wärmepumpe (Kompaktanlage) inkl. Montage\n"
        "✓ Fußbodenheizung im ganzen Haus\n"
        "✓ Be- und Entlüftungsanlage mit Wärmerückgewinnung\n"
        "✓ Elektroinstallation & Sanitärrohinstallation\n"
        "✓ Holztreppe EG/DG\n"
        "✓ Technikpaket für KfW 55 inkl. Blower-Door-Test\n\n"
        "Das Ausbauhaus von massa haus bedeutet:\n"
        "Wetterfeste Hülle in Holzständerbauweise, geplant (Architektenleistungen "
        "enthalten) und aufgestellt – Sie bauen im eigenen Tempo aus."
    )

    preisblock = (
        f"Grundstück:      {_fmt(gs)}\n"
        f"Ausbauhaus:      {_fmt(ab)}\n"
        f"Technikpaket:    {_fmt(tp)}\n"
        f"─────────────────────────────────\n"
        f"Gesamt:          {_fmt(gesamt)}"
    )

    grundstueck_info = f"Grundstücksgröße: {groesse} m²\n" if groesse else ""

    main_teile = [
        kerntext,
        "",
        "── PREISÜBERSICHT ──",
        preisblock,
        "",
        "── GRUNDSTÜCK ──",
        grundstueck_info + f"Ort: {ort}",
    ]
    if anbieter:
        main_teile += ["", f"Anbieter: {anbieter}"]
    main_teile += [
        "",
        "── HINWEIS ──",
        "Alle Preisangaben inkl. 19 % MwSt. | Hauspreise massa haus Stand 11/2025.",
        "Grundstücksangaben gemäß Anbieter. Irrtümer vorbehalten.",
    ]

    lage_note = _lage_beschreibung(plz, ort)
    return "\n".join(main_teile), lage_note


# ── Propstack API ─────────────────────────────────────────────

def _post_to_propstack(payload: dict):
    try:
        r = requests.post(
            f"{PROPSTACK_BASE}/units",
            headers=HEADERS,
            json=payload,
            timeout=15,
        )
        if r.status_code in (200, 201):
            data = r.json()
            unit_id = data.get("id") or data.get("unit", {}).get("id", "")
            return True, str(unit_id)
        else:
            return False, f"HTTP {r.status_code}: {r.text[:300]}"
    except Exception as e:
        return False, str(e)


def _lade_bilder_fuer_haustyp(haustyp: str) -> list:
    """
    Sucht Bilder + Grundriss für einen Haustyp im Bild-Ordner.
    Gibt Liste von (pfad, titel, is_grundriss) zurück.
    """
    import glob

    kurz = {
        "LifeStyle": "LS", "ComfortStyle": "CS", "FamilyStyle": "FS",
        "TwinStyle": "TS", "TrendStyle": "TR", "Trend": "TR",
        "UnitStyle": "US", "UrbanStyle": "UrS",
    }
    m = re.match(r"(\w+)\s+([\d\.]+)\s*([A-Z\.]+)", haustyp)
    if not m:
        return []
    reihe, nummer, variant = m.group(1), m.group(2), m.group(3).replace(".", "")
    prefix = kurz.get(reihe, reihe[:2])
    num_clean = nummer.replace(".", "-").rstrip("-")
    datei_prefix = f"{prefix}-{num_clean}-{variant}"

    bilder = []
    for unterordner in os.listdir(IMG_BASISORDNER):
        ordner_pfad = os.path.join(IMG_BASISORDNER, unterordner)
        if not os.path.isdir(ordner_pfad):
            continue
        for muster in [f"{datei_prefix}-bild*.jpg", f"{datei_prefix}-bild*.png",
                        f"{datei_prefix.lower()}-bild*.jpg"]:
            for pfad in sorted(glob.glob(os.path.join(ordner_pfad, muster))):
                bilder.append((pfad, "Hausansicht", False))
        for muster in [f"{datei_prefix}-grundriss.jpg", f"{datei_prefix}-grundriss.png"]:
            for pfad in glob.glob(os.path.join(ordner_pfad, muster)):
                bilder.append((pfad, "Grundriss", True))

    return bilder


def _upload_bilder(property_id: str, bilder: list):
    """Lädt Bilder als Dokumente in Propstack hoch."""
    for pfad, titel, _ in bilder:
        dateiname = os.path.basename(pfad)
        try:
            with open(pfad, "rb") as f:
                r = requests.post(
                    f"{PROPSTACK_BASE}/documents",
                    headers=HEADERS_UPLOAD,
                    files={"document[file]": (dateiname, f, "image/jpeg")},
                    data={"document[property_id]": property_id,
                          "document[title]": titel,
                          "document[is_private]": "false"},
                    timeout=30,
                )
            if r.status_code != 200:
                print(f"      ⚠️  Bild-Upload fehlgeschlagen: {dateiname} ({r.status_code})")
        except Exception as e:
            print(f"      ⚠️  Bild-Upload Fehler: {dateiname}: {e}")


def _payload(titel: str, beschr_main: str, beschr_lage: str, plz: str, ort: str,
             preis_gesamt: float) -> dict:
    payload = {
        "property": {
            "object_type":      "LIVING",   # LIVING (nicht LAND) – sonst leere Detailseite!
            "rs_type":          "PLOT",
            "marketing_type":   "BUY",
            "title":            titel,
            "zip_code":         str(plz) if plz else "",
            "city":             str(ort) if ort else "",
            "description_note": beschr_main,
            "location_note":    beschr_lage,
            "hide_address":     True,       # Blindadresse – keine Straße!
        }
    }
    if preis_gesamt and preis_gesamt > 0:
        payload["property"]["price"] = preis_gesamt
    return payload


# ── Hauptfunktion ─────────────────────────────────────────────

def importiere_in_propstack():
    print("=" * 60)
    print("  PROPSTACK IMPORT – massa haus Sophia")
    print("=" * 60)

    if not os.path.exists(EXCEL_OUT):
        print(f"❌ Excel nicht gefunden: {EXCEL_OUT}")
        return

    wb = load_workbook(EXCEL_OUT)
    ws = wb.active

    to_import = []
    for row_idx in range(2, ws.max_row + 1):
        status_val = ws.cell(row_idx, COL["status"]).value
        if status_val and str(status_val).strip().lower() in IMPORT_STATI:
            to_import.append(row_idx)

    if not to_import:
        print("ℹ️  Keine Zeilen mit Status 'Import' gefunden.")
        return

    print(f"\n📋 {len(to_import)} Grundstück(e) zum Import\n")
    erfolge = 0

    for row_idx in to_import:
        def cell(col_name):
            v = ws.cell(row_idx, COL[col_name]).value
            return str(v).strip() if v else ""

        plz      = cell("plz")
        ort      = cell("ort")
        groesse  = cell("groesse")
        haustyp  = cell("haustyp")
        anbieter = cell("anbieter")
        gs_preis = _parse_grundstueck_preis(cell("preis"))

        print(f"  📍 {ort} ({plz}) | {groesse} | {gs_preis:,.0f} € | {haustyp}")

        if not haustyp:
            print("  ⚠️  Kein Haustyp eingetragen – übersprungen.")
            ws.cell(row=row_idx, column=COL["status"]).value = "Import fehlt Haustyp"
            continue

        hauspreise = _lade_hauspreise(haustyp)
        if not hauspreise:
            ws.cell(row=row_idx, column=COL["status"]).value = "Import fehlt Haustyp"
            continue

        gesamt = gs_preis + hauspreise["ausbauhaus"] + hauspreise["technikpaket"]
        titel = _titel(ort, haustyp)
        beschr_main, beschr_lage = _beschreibung(
            gs_preis, hauspreise, groesse, plz, ort, haustyp, anbieter
        )
        payload = _payload(titel, beschr_main, beschr_lage, plz, ort, gesamt)
        payload["property"]["broker_id"] = BROKER_ID

        bilder = _lade_bilder_fuer_haustyp(haustyp)
        print(f"  🖼  {len(bilder)} Bilder gefunden für {haustyp}")

        ok, result = _post_to_propstack(payload)
        if ok:
            print(f"  ✅ Angelegt: {titel[:55]}...")
            print(f"     Gesamtpreis: {_fmt(gesamt)} | ID: {result}")
            if bilder:
                _upload_bilder(result, bilder)
                print(f"     🖼  {len(bilder)} Bilder hochgeladen")
            ws.cell(row=row_idx, column=COL["status"]).value       = "In Propstack"
            ws.cell(row=row_idx, column=COL["in_propstack"]).value = result
            erfolge += 1
        else:
            print(f"  ❌ Fehler: {result}")

    wb.save(EXCEL_OUT)
    print(f"\n{'=' * 60}")
    print(f"  ✅ {erfolge} von {len(to_import)} Entwürfe angelegt")
    print(f"  Excel gespeichert: {EXCEL_OUT}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    importiere_in_propstack()
